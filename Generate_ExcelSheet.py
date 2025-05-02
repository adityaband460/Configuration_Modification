from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3


# def createConfigurationFile():
#     pass

# def UpdateDatabaseFile():
#     pass

# openpyxl indexing starts from 1
wb = Workbook()
worksheet = wb.active
worksheet.title = "Sheet1"

ROW_TITLES = ['','PLMN', 'TAC', 'CELL_IDENTITY','PCI', 'DL_ARFCN_',
              'UL_ARFCN', 'BAND_INDICATOR','MME_IP_ADDRESS', 'BBU_IP_ADDRESS']

COL_TITLES = ['','System','Sector-0', 'Sector-1', 'Sector-2']

for row_num, title in enumerate(ROW_TITLES,start=1):
    worksheet.cell(row_num, 1, title)

for col_num, title in enumerate(COL_TITLES,start=1):
    worksheet.cell(1, col_num, title)


db_path = "./20250314_amannagar_b1_zte6-RelocTimers-alarm2.sqlite"

with sqlite3.connect(db_path) as conn:
    cursor = conn.cursor()
    # Get PLMN 
    get_plmn_PLMNTable = """ SELECT PLMNID FROM PLMNTable """
    cursor.execute(get_plmn_PLMNTable)

    PLMN = int(cursor.fetchone()[0])
    worksheet.cell(row=2,column= 2, value=PLMN)

    # Get TAC
    # EpcTable --> TAC
    get_tac_EpcTable = """ SELECT TAC FROM EpcTable """
    cursor.execute(get_tac_EpcTable)

    TAC = int(cursor.fetchone()[0])
    worksheet.cell(row=3,column= 2, value=TAC)

    # Get CELL_IDENTITY
    get_cell_idnetity_0 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '0' """
    get_cell_identity_1 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '1' """
    get_cell_identity_2 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '2' """
    
    cursor.execute(get_cell_idnetity_0)
    CELL_IDENTITY_0 = int(cursor.fetchone()[0])
    worksheet.cell(row=4,column= 3, value=CELL_IDENTITY_0)
    
    cursor.execute(get_cell_identity_1)
    CELL_IDENTITY_1 = int(cursor.fetchone()[0])
    worksheet.cell(row=4,column= 4, value=CELL_IDENTITY_1)
    
    cursor.execute(get_cell_identity_2)
    CELL_IDENTITY_2 = int(cursor.fetchone()[0])
    worksheet.cell(row=4,column= 5, value=CELL_IDENTITY_2)

    # Get PCI
    # UPDATE RFParamsTable SET PhyCellID = ? 
    # WHERE CellIndex = ? """
    get_pci_0 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '0' """
    get_pci_1 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '1' """
    get_pci_2 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '2' """

    cursor.execute(get_pci_0)
    PCI_0 = int(cursor.fetchone()[0])
    worksheet.cell(row=5,column= 3, value=PCI_0)

    cursor.execute(get_pci_1)
    PCI_1 = int(cursor.fetchone()[0])
    worksheet.cell(row=5,column= 4, value=PCI_1)
    
    cursor.execute(get_pci_2)
    PCI_2 = int(cursor.fetchone()[0])
    worksheet.cell(row=5,column= 5, value=PCI_2)
    
    # Get DL_ARFCN
    # update_earfcn_dl_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNDL = ? 
    # WHERE CellIndex = ? """

    get_dl_arfcn_0 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '0' """
    cursor.execute(get_dl_arfcn_0)
    DL_ARFCN_0 = int(cursor.fetchone()[0])
    worksheet.cell(row=6,column= 3, value=DL_ARFCN_0)

    get_dl_arfcn_1 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '1' """
    cursor.execute(get_dl_arfcn_1)
    DL_ARFCN_1 = int(cursor.fetchone()[0])
    worksheet.cell(row=6,column= 4, value=DL_ARFCN_1)

    get_dl_arfcn_2 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '2' """
    cursor.execute(get_dl_arfcn_2)
    DL_ARFCN_2 = int(cursor.fetchone()[0])
    worksheet.cell(row=6,column= 5, value=DL_ARFCN_2)

    # Get UL_ARFCN
    # update_earfcn_ul_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNUL = ?
    # WHERE CellIndex = ? """

    get_ul_arfcn_0 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '0' """
    cursor.execute(get_ul_arfcn_0)
    UL_ARFCN_0 = int(cursor.fetchone()[0])
    worksheet.cell(row=7,column= 3, value=UL_ARFCN_0)

    get_ul_arfcn_1 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '1' """
    cursor.execute(get_ul_arfcn_1)
    UL_ARFCN_1 = int(cursor.fetchone()[0])
    worksheet.cell(row=7,column= 4, value=UL_ARFCN_1)

    get_ul_arfcn_2 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '2' """
    cursor.execute(get_ul_arfcn_2)
    UL_ARFCN_2 = int(cursor.fetchone()[0])
    worksheet.cell(row=7,column= 5, value=UL_ARFCN_2)

    # Get BAND_INDICATOR
    # update_band_indicator_query_RFParamsTable = """ UPDATE RFParamsTable SET FreqBandIndicator = ?
    # WHERE CellIndex = ? """

    get_band_indicator_0 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '0' """
    cursor.execute(get_band_indicator_0)
    BAND_INDICATOR_0 = int(cursor.fetchone()[0])
    worksheet.cell(row=8,column= 3, value=BAND_INDICATOR_0)

    get_band_indicator_1 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '1' """
    cursor.execute(get_band_indicator_1)
    BAND_INDICATOR_1 = int(cursor.fetchone()[0])
    worksheet.cell(row=8,column= 4, value=BAND_INDICATOR_1)

    get_band_indicator_2 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '2' """
    cursor.execute(get_band_indicator_2)
    BAND_INDICATOR_2 = int(cursor.fetchone()[0])
    worksheet.cell(row=8,column= 5, value=BAND_INDICATOR_2)

    # Get MME_IP_ADDRESS
    get_mme_ip_address = """ SELECT S1SigLinkServerList FROM MMECommInfoTable """
    cursor.execute(get_mme_ip_address)
    MME_IP_ADDRESS = cursor.fetchone()[0]
    worksheet.cell(row=9,column= 2, value=MME_IP_ADDRESS)

    # Get BBU_IP_ADDRESS
    # UPDATE globalEnbIdInfoTable SET henb_self_address 
    get_bbu_ip_address = """ SELECT henb_self_address FROM globalEnbIdInfoTable """
    cursor.execute(get_bbu_ip_address)
    BBU_IP_ADDRESS = cursor.fetchone()[0]
    worksheet.cell(row=10,column= 2, value=BBU_IP_ADDRESS)




# just for fomatting 
for i in range(1,101):
    worksheet.row_dimensions[i].height = 20
    worksheet.column_dimensions[get_column_letter(i)].width = 20

# Save the workbook
wb.save("sqlite_db.xlsx")