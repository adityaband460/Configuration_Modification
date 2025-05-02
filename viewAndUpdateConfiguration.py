from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3

ROW_TITLES = [
    '', 'PLMN', 'TAC', 'CELL_IDENTITY', 'PCI', 'DL_ARFCN_',
    'UL_ARFCN', 'BAND_INDICATOR', 'MME_IP_ADDRESS', 'BBU_IP_ADDRESS'
]

COL_TITLES = ['', 'System', 'Sector-0', 'Sector-1', 'Sector-2']


def createConfigurationFile(db_path: str, output_path: str = "sqlite_db.xlsx"):
    # Set up workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write row and column headers
    for row_num, title in enumerate(ROW_TITLES, start=1):
        ws.cell(row=row_num, column=1, value=title)
    for col_num, title in enumerate(COL_TITLES, start=1):
        ws.cell(row=1, column=col_num, value=title)

    # Connect to SQLite and populate cells
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()

        # PLMN
        cur.execute("SELECT PLMNID FROM PLMNTable")
        ws.cell(row=2, column=2, value=int(cur.fetchone()[0]))

        # TAC
        cur.execute("SELECT TAC FROM EpcTable")
        ws.cell(row=3, column=2, value=int(cur.fetchone()[0]))

        # CELL_IDENTITY for sectors 0–2
        for idx in range(3):
            cur.execute(
                "SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = ?",
                (str(idx),)
            )
            ws.cell(row=4, column=3+idx, value=int(cur.fetchone()[0]))

        # PCI for sectors 0–2
        for idx in range(3):
            cur.execute(
                "SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = ?",
                (str(idx),)
            )
            ws.cell(row=5, column=3+idx, value=int(cur.fetchone()[0]))

        # DL_ARFCN for sectors 0–2
        for idx in range(3):
            cur.execute(
                "SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = ?",
                (str(idx),)
            )
            ws.cell(row=6, column=3+idx, value=int(cur.fetchone()[0]))

        # UL_ARFCN for sectors 0–2
        for idx in range(3):
            cur.execute(
                "SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = ?",
                (str(idx),)
            )
            ws.cell(row=7, column=3+idx, value=int(cur.fetchone()[0]))

        # BAND_INDICATOR for sectors 0–2
        for idx in range(3):
            cur.execute(
                "SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = ?",
                (str(idx),)
            )
            ws.cell(row=8, column=3+idx, value=int(cur.fetchone()[0]))

        # MME_IP_ADDRESS
        cur.execute("SELECT S1SigLinkServerList FROM MMECommInfoTable")
        ws.cell(row=9, column=2, value=cur.fetchone()[0])

        # BBU_IP_ADDRESS
        cur.execute("SELECT henb_self_address FROM globalEnbIdInfoTable")
        ws.cell(row=10, column=2, value=cur.fetchone()[0])

    # Formatting: row heights and column widths
    for i in range(1, 51):
        ws.row_dimensions[i].height = 20
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Save the workbook
    wb.save(output_path)
    print(f"Configuration file written to '{output_path}'.")


def updateDataBaseFile(db_path: str, conf_file: str = "sqlite_db.xlsx"):
    # copy content from conf_file to db_path
    pass


def main():
    db_path = "./20250314_amannagar_b1_zte6-RelocTimers-alarm2.sqlite"
    createConfigurationFile(db_path)


if __name__ == "__main__":
    main()
