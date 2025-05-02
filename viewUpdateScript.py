from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sqlite3
import sys


db_path = "./20250314_amannagar_b1_zte6-RelocTimers-alarm2.sqlite"

def createCsvFromDatabaseFile(db_path: str, output_path: str = "sqlite_db.xlsx"):
    wb = Workbook()
    worksheet = wb.active
    worksheet.title = "Sheet1"

    ROW_TITLES = ['','PLMN', 'TAC', 'CELL_IDENTITY','PCI', 'DL_ARFCN_',
                'UL_ARFCN', 'BAND_INDICATOR','MME_IP_ADDRESS', 'BBU_IP_ADDRESS']

    COL_TITLES = ['','System','Sector-0', 'Sector-1', 'Sector-2']

    for row_num, title in enumerate(ROW_TITLES,start=1):
        worksheet.cell(row_num, 1, title)

    for col_num, title in enumerate(COL_TITLES,start=1):
        worksheet.cell(1, col_num, title)

    try:    
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            # Get PLMN 
            get_plmn_PLMNTable = """ SELECT PLMNID FROM PLMNTable """
            cursor.execute(get_plmn_PLMNTable)

            PLMN = int(cursor.fetchone()[0])
            worksheet.cell(row=2,column= 2, value=PLMN)

            # Get TAC
            # EpcTable --> TAC
            get_tac_EpcTable = """ SELECT TAC FROM EpcTable """
            cursor.execute(get_tac_EpcTable)

            TAC = int(cursor.fetchone()[0])
            worksheet.cell(row=3,column= 2, value=TAC)

            # Get CELL_IDENTITY
            get_cell_idnetity_0 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '0' """
            get_cell_identity_1 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '1' """
            get_cell_identity_2 = """ SELECT CellIdentity FROM CellIdentityTable WHERE CellIndex = '2' """
            
            cursor.execute(get_cell_idnetity_0)
            CELL_IDENTITY_0 = int(cursor.fetchone()[0])
            worksheet.cell(row=4,column= 3, value=CELL_IDENTITY_0)
            
            cursor.execute(get_cell_identity_1)
            CELL_IDENTITY_1 = int(cursor.fetchone()[0])
            worksheet.cell(row=4,column= 4, value=CELL_IDENTITY_1)
            
            cursor.execute(get_cell_identity_2)
            CELL_IDENTITY_2 = int(cursor.fetchone()[0])
            worksheet.cell(row=4,column= 5, value=CELL_IDENTITY_2)

            # Get PCI
            # UPDATE RFParamsTable SET PhyCellID = ? 
            # WHERE CellIndex = ? """
            get_pci_0 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '0' """
            get_pci_1 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '1' """
            get_pci_2 = """ SELECT PhyCellID FROM RFParamsTable WHERE CellIndex = '2' """

            cursor.execute(get_pci_0)
            PCI_0 = int(cursor.fetchone()[0])
            worksheet.cell(row=5,column= 3, value=PCI_0)

            cursor.execute(get_pci_1)
            PCI_1 = int(cursor.fetchone()[0])
            worksheet.cell(row=5,column= 4, value=PCI_1)
            
            cursor.execute(get_pci_2)
            PCI_2 = int(cursor.fetchone()[0])
            worksheet.cell(row=5,column= 5, value=PCI_2)
            
            # Get DL_ARFCN
            # update_earfcn_dl_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNDL = ? 
            # WHERE CellIndex = ? """

            get_dl_arfcn_0 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '0' """
            cursor.execute(get_dl_arfcn_0)
            DL_ARFCN_0 = int(cursor.fetchone()[0])
            worksheet.cell(row=6,column= 3, value=DL_ARFCN_0)

            get_dl_arfcn_1 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '1' """
            cursor.execute(get_dl_arfcn_1)
            DL_ARFCN_1 = int(cursor.fetchone()[0])
            worksheet.cell(row=6,column= 4, value=DL_ARFCN_1)

            get_dl_arfcn_2 = """ SELECT EARFCNDL FROM RFParamsTable WHERE CellIndex = '2' """
            cursor.execute(get_dl_arfcn_2)
            DL_ARFCN_2 = int(cursor.fetchone()[0])
            worksheet.cell(row=6,column= 5, value=DL_ARFCN_2)

            # Get UL_ARFCN
            # update_earfcn_ul_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNUL = ?
            # WHERE CellIndex = ? """

            get_ul_arfcn_0 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '0' """
            cursor.execute(get_ul_arfcn_0)
            UL_ARFCN_0 = int(cursor.fetchone()[0])
            worksheet.cell(row=7,column= 3, value=UL_ARFCN_0)

            get_ul_arfcn_1 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '1' """
            cursor.execute(get_ul_arfcn_1)
            UL_ARFCN_1 = int(cursor.fetchone()[0])
            worksheet.cell(row=7,column= 4, value=UL_ARFCN_1)

            get_ul_arfcn_2 = """ SELECT EARFCNUL FROM RFParamsTable WHERE CellIndex = '2' """
            cursor.execute(get_ul_arfcn_2)
            UL_ARFCN_2 = int(cursor.fetchone()[0])
            worksheet.cell(row=7,column= 5, value=UL_ARFCN_2)

            # Get BAND_INDICATOR
            # update_band_indicator_query_RFParamsTable = """ UPDATE RFParamsTable SET FreqBandIndicator = ?
            # WHERE CellIndex = ? """

            get_band_indicator_0 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '0' """
            cursor.execute(get_band_indicator_0)
            BAND_INDICATOR_0 = int(cursor.fetchone()[0])
            worksheet.cell(row=8,column= 3, value=BAND_INDICATOR_0)

            get_band_indicator_1 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '1' """
            cursor.execute(get_band_indicator_1)
            BAND_INDICATOR_1 = int(cursor.fetchone()[0])
            worksheet.cell(row=8,column= 4, value=BAND_INDICATOR_1)

            get_band_indicator_2 = """ SELECT FreqBandIndicator FROM RFParamsTable WHERE CellIndex = '2' """
            cursor.execute(get_band_indicator_2)
            BAND_INDICATOR_2 = int(cursor.fetchone()[0])
            worksheet.cell(row=8,column= 5, value=BAND_INDICATOR_2)

            # Get MME_IP_ADDRESS
            get_mme_ip_address = """ SELECT S1SigLinkServerList FROM MMECommInfoTable """
            cursor.execute(get_mme_ip_address)
            MME_IP_ADDRESS = cursor.fetchone()[0]
            worksheet.cell(row=9,column= 2, value=MME_IP_ADDRESS)

            # Get BBU_IP_ADDRESS
            # UPDATE globalEnbIdInfoTable SET henb_self_address 
            get_bbu_ip_address = """ SELECT henb_self_address FROM globalEnbIdInfoTable """
            cursor.execute(get_bbu_ip_address)
            BBU_IP_ADDRESS = cursor.fetchone()[0]
            worksheet.cell(row=10,column= 2, value=BBU_IP_ADDRESS)
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
        exit(1)



    # just for fomatting 
    for i in range(1,101):
        worksheet.row_dimensions[i].height = 20
        worksheet.column_dimensions[get_column_letter(i)].width = 20

    # Save the workbook
    wb.save(output_path)
    print(f"Excel file '{output_path}' created successfully.")

def updateDatabaseFileFromCsv(db_path: str, input_path: str = "sqlite_db.xlsx"):
    
    
    file_path = input_path

    try:
        workbook = load_workbook(filename = file_path)
    except FileNotFoundError:
        print(f"Excel file not found: {file_path} Generate it by giving the command 'view'")
        exit(1)
        
        
    sheet = workbook['Sheet1']

    PLMN = sheet['B2'].value
    TAC = sheet['B3'].value

    CELL_IDENTITY_0 = sheet['C4'].value
    CELL_IDENTITY_1 = sheet['D4'].value
    CELL_IDENTITY_2 = sheet['E4'].value

    PCI_0 = sheet['C5'].value
    PCI_1 = sheet['D5'].value
    PCI_2 = sheet['E5'].value

    DL_ARFCN_0 = sheet['C6'].value
    DL_ARFCN_1 = sheet['D6'].value
    DL_ARFCN_2 = sheet['E6'].value

    UL_ARFCN_0 = sheet['C7'].value
    UL_ARFCN_1 = sheet['D7'].value
    UL_ARFCN_2 = sheet['E7'].value

    BAND_INDICATOR_0 = sheet['C8'].value
    BAND_INDICATOR_1 = sheet['D8'].value
    BAND_INDICATOR_2 = sheet['E8'].value

    MME_IP_ADDRESS = sheet['B9'].value
    BBU_IP_ADDRESS = sheet['B10'].value



    # print(PLMN, TAC)

    db_path = "./20250314_amannagar_b1_zte6-RelocTimers-alarm2.sqlite"
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()

            # Update PLMN
            update_plmn_query_PLMNTable = """ UPDATE PLMNTable SET PLMNID = ? """
            update_plmn_query_LTENeighbourCellInfoTable = """ UPDATE LTENeighbourCellInfoTable SET PLMNID = ? """ 
            update_plmn_query_PeerEnbCommInfoTable = """ UPDATE PeerEnbCommInfoTable SET PLMNID = ? """ 
            update_plmn_query_OperatorInfoTable = """ UPDATE OperatorInfoTable SET plmnId = ? """ 
            update_plmn_query_UTRANNeighbourCellInfoTable = """ UPDATE UTRANNeighbourCellInfoTable SET PLMNID = ? """ 
            update_plmn_query_GERANNeighbourCellInfoTable = """ UPDATE GERANNeighbourCellInfoTable SET PLMNID = ? """ 


            cursor.execute(update_plmn_query_PLMNTable, (PLMN,))
            cursor.execute(update_plmn_query_LTENeighbourCellInfoTable, (PLMN,))
            cursor.execute(update_plmn_query_PeerEnbCommInfoTable, (PLMN,))
            cursor.execute(update_plmn_query_OperatorInfoTable, (PLMN,))
            cursor.execute(update_plmn_query_UTRANNeighbourCellInfoTable, (PLMN,))
            cursor.execute(update_plmn_query_GERANNeighbourCellInfoTable, (PLMN,))


            # Update TAC
            # LTENeighbourCellInfoTable --> X_VENDOR_TAC
            # PeerEnbCommInfoTable --> TAC
            # EpcTable --> TAC

            update_tac_query_LTENeighbourCellInfoTable = """ UPDATE LTENeighbourCellInfoTable SET X_VENDOR_TAC = ? """
            update_tac_query_PeerEnbCommInfoTable = """ UPDATE PeerEnbCommInfoTable SET TAC = ? """
            update_tac_query_EpcTable = """ UPDATE EpcTable SET TAC = ? """

            cursor.execute(update_tac_query_LTENeighbourCellInfoTable, (TAC,))
            cursor.execute(update_tac_query_PeerEnbCommInfoTable, (TAC,))
            cursor.execute(update_tac_query_EpcTable, (TAC,))

            # Update Cell Identity
            # CellIdentityTable --> CellIdentity

            update_cell_identity_query_CellIdentityTable = """ UPDATE CellIdentityTable SET CellIdentity = ? 
            WHERE CellIndex = ? """

            cursor.execute(update_cell_identity_query_CellIdentityTable, (CELL_IDENTITY_0,'0'))
            cursor.execute(update_cell_identity_query_CellIdentityTable, (CELL_IDENTITY_1,'1'))
            cursor.execute(update_cell_identity_query_CellIdentityTable, (CELL_IDENTITY_2,'2'))

            # Update PCI
            # RFParamsTable --> PhyCellID 

            update_pci_query_RFParamsTable = """ UPDATE RFParamsTable SET PhyCellID = ? 
            WHERE CellIndex = ? """

            cursor.execute(update_pci_query_RFParamsTable, (PCI_0,'0'))
            cursor.execute(update_pci_query_RFParamsTable, (PCI_1,'1'))
            cursor.execute(update_pci_query_RFParamsTable, (PCI_2,'2'))

            # Update EARFCNDL
            # RFParamsTable    --> EARFCNDL

            update_earfcn_dl_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNDL = ? 
            WHERE CellIndex = ? """

            cursor.execute(update_earfcn_dl_query_RFParamsTable, (DL_ARFCN_0,'0'))
            cursor.execute(update_earfcn_dl_query_RFParamsTable, (DL_ARFCN_1,'1'))
            cursor.execute(update_earfcn_dl_query_RFParamsTable, (DL_ARFCN_2,'2'))

            # Update EARFCNUL
            # RFParamsTable    --> EARFCNUL

            update_earfcn_ul_query_RFParamsTable = """ UPDATE RFParamsTable SET EARFCNUL = ?
            WHERE CellIndex = ? """

            cursor.execute(update_earfcn_ul_query_RFParamsTable, (UL_ARFCN_0,'0'))
            cursor.execute(update_earfcn_ul_query_RFParamsTable, (UL_ARFCN_1,'1'))
            cursor.execute(update_earfcn_ul_query_RFParamsTable, (UL_ARFCN_2,'2'))


            # Update Band Indicator
            # RFParamsTable --- > FreqBandIndicator

            update_band_indicator_query_RFParamsTable = """ UPDATE RFParamsTable SET FreqBandIndicator = ?
            WHERE CellIndex = ? """

            cursor.execute(update_band_indicator_query_RFParamsTable, (BAND_INDICATOR_0,'0'))
            cursor.execute(update_band_indicator_query_RFParamsTable, (BAND_INDICATOR_1,'1'))
            cursor.execute(update_band_indicator_query_RFParamsTable, (BAND_INDICATOR_2,'2'))

            # Update MME IP Address
            # MMECommInfoTable --> S1SigLinkServerList

            update_mme_ip_query_MMECommInfoTable = """ UPDATE MMECommInfoTable SET S1SigLinkServerList = ?"""

            cursor.execute(update_mme_ip_query_MMECommInfoTable, (MME_IP_ADDRESS,))


            # Update BBU IP Address
            # IpInterfaceTable     ---> IPInterfaceIPAddress where X_VENDOR_INTERFACE_TYPE ,X_VENDOR_SOURCE_PORT_NUMBER
            # 'OAM_S1AP_INTERFACE','36412'
            # 'OAM_X2AP_INTERFACE',36422
            # 'OAM_GTPU_INTERFACE',2152, 
            # globalEnbIdInfoTable  --> henb_self_address



            # Tr69InitParamsTable  --> ConnectionRequestURL --> 'http://10.131.183.31:15210'
            # EnodeBEMSIpsecTable   --> ENBIpAddress
            # ipAddressTable        --> ?

            update_bbu_ip_query_IpInterfaceTable = """ UPDATE IpInterfaceTable SET IPInterfaceIPAddress = ? 
            where Enable = '1' and (
                (X_VENDOR_INTERFACE_TYPE = 'OAM_S1AP_INTERFACE' and X_VENDOR_SOURCE_PORT_NUMBER = '36412') or 
                (X_VENDOR_INTERFACE_TYPE = 'OAM_X2AP_INTERFACE' and X_VENDOR_SOURCE_PORT_NUMBER = 36422) or
                (X_VENDOR_INTERFACE_TYPE = 'OAM_GTPU_INTERFACE' and X_VENDOR_SOURCE_PORT_NUMBER = 2152)) """
            
            update_bbu_ip_query_globalEnbIdInfoTable = """ UPDATE globalEnbIdInfoTable SET henb_self_address = ? """
            

            cursor.execute(update_bbu_ip_query_IpInterfaceTable, (BBU_IP_ADDRESS,))
            cursor.execute(update_bbu_ip_query_globalEnbIdInfoTable, (BBU_IP_ADDRESS,))
            # cursor.execute("UPDATE IpInterfaceTable SET IPInterfaceIPAddress = ? WHERE Alias = 'cpe-s1ap-1'", (BBU_IP_ADDRESS,))

            
            # commit the changes to the database
            conn.commit()
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
        exit(1)
    
    print(f"Database file '{db_path}' updated successfully.")


def main():
    command = None
    if len(sys.argv) < 2:
        command = "view"
    
    command = sys.argv[1].lower()

    if command == "view":
        createCsvFromDatabaseFile(db_path)
    elif command == "update":
        updateDatabaseFileFromCsv(db_path)

if __name__ == "__main__":
    main()